import re
import time
import traceback
from datetime import datetime
from pathlib import Path
import pandas as pd
import fitz 
import requests
from loguru import logger
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "MTE"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")

URL_MTE = "https://eprocesso.sit.trabalho.gov.br/Entrar?ReturnUrl=%2FCertidao%2FEmitir"
TIMEOUT = 40_000
REGEX_VALIDADE = r"Válida até:\s*(\d{2}/\d{2}/\d{4})"

# === Funções utilitárias ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

# === Processo principal ===
def processar_mte():
    logger.add("execucao_mte.log", rotation="1 MB")
    logger.info("Iniciando automação GOV.BR no MTE via certificado digital")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.connect_over_cdp("http://localhost:9222")
        context = browser.contexts[0]
        page = context.new_page()
        page.bring_to_front()

        try:
            logger.info(f"Acessando {URL_MTE}...")
            page.goto(URL_MTE, timeout=TIMEOUT)

            # Clica em "Entrar com GOV.BR"
            logger.info("Clicando em 'Entrar com GOV.BR'")
            page.click("#janela-login-gov-br a")

            # Clica no botão "Seu certificado digital"
            logger.info("Selecionando login com certificado digital...")
            page.locator("#login_certificate").click()

            # Aguarda aparecer a lista de certificados
            logger.info("Aguardando lista de certificados...")
            certificados = page.locator("div[role='option'], div.certificado, div.card-certificado").first
            certificados.wait_for(state="visible", timeout=20000)

            # Seleciona o primeiro certificado
            logger.info("Selecionando primeiro certificado...")
            certificados.click()

            logger.success("Login via certificado digital concluído!")
            time.sleep(10)  # dá tempo de redirecionar e carregar

        except Exception as e:
            motivo = f"{type(e)._name_}: {e}"
            logger.error(f"Erro durante processo de login: {motivo}")
            traceback.print_exc()

        finally:
            context.close()
            browser.close()

    logger.info("Processo concluído.")


# === Execução ===
if __name__ == "__main__":
    processar_mte()
