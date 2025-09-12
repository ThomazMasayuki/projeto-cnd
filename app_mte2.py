import re
import time
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import fitz  # PyMuPDF
import requests
from loguru import logger
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "MTE"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")

URL_MTE = "https://eprocesso.sit.trabalho.gov.br/Entrar?ReturnUrl=%2FCertidao%2FEmitir"
TIMEOUT = 40_000
REGEX_VALIDADE = r"Válida até:\s*(\d{2}/\d{2}/\d{4})"

# === GOV.BR login ===
CPF_LOGIN = "02448982198"     # coloque aqui seu CPF
SENHA_LOGIN = "15468973Tt*"   # coloque aqui sua senha GOV.BR

# === 2Captcha (hCaptcha) ===
API_KEY_2CAPTCHA = "30924a201f06a3b554d7479c487fee8e"  # sua chave real do 2Captcha
SITEKEY = "93b08d40-d46c-400a-ba07-6f91cda815b9"       # encontrado no HTML
PAGEURL = "https://sso.acesso.gov.br/login"

# === Funções utilitárias ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

# === 2Captcha: hCaptcha ===
def solicitar_hcaptcha(api_key, sitekey, url):
    payload = {
        "key": api_key,
        "method": "hcaptcha",
        "sitekey": sitekey,
        "pageurl": url,
        "json": 1
    }
    r = requests.post("http://2captcha.com/in.php", data=payload)
    return r.json().get("request")

def obter_resultado(api_key, captcha_id, tentativas=30, intervalo=7):
    for _ in range(tentativas):
        time.sleep(intervalo)
        r = requests.get("http://2captcha.com/res.php", params={
            "key": api_key,
            "action": "get",
            "id": captcha_id,
            "json": 1
        })
        data = r.json()
        if data.get("status") == 1:
            return data.get("request")
        if data.get("request") != "CAPCHA_NOT_READY":
            raise RuntimeError(f"Erro no 2Captcha: {data}")
    raise TimeoutError("Captcha não resolvido dentro do tempo.")

# === Processo principal ===
def processar_mte():
    logger.add("execucao_mte.log", rotation="1 MB")
    logger.info("Iniciando automação GOV.BR no MTE")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/114.0.0.0 Safari/537.36",
            viewport={"width": 1920, "height": 1080})
        page = context.new_page()

        try:
            logger.info(f"Acessando {URL_MTE}...")
            page.goto(URL_MTE, timeout=TIMEOUT)

            # Clica em "Entrar com GOV.BR"
            page.click("#janela-login-gov-br a")

            # Preenche CPF
            logger.info("Preenchendo CPF...")
            campo_cpf = page.get_by_role("textbox", name="Digite seu CPF")
            campo_cpf.fill("")
            campo_cpf.type(CPF_LOGIN, delay=150)
            time.sleep(5)
            
            # Clica no botão continuar
            logger.info("Clicando no botão 'Continuar'")
            btn = page.locator("button#enter-account-id")
            btn.click()

            # Solicita resolução do hCaptcha
            logger.info("Solicitando resolução do hCaptcha via 2Captcha...")
            captcha_id = solicitar_hcaptcha(API_KEY_2CAPTCHA, SITEKEY, PAGEURL)
            token = obter_resultado(API_KEY_2CAPTCHA, captcha_id)
            logger.success("Token do hCaptcha resolvido!")

            # Injetar token resolvido no campo correto
            # Aguarda o campo hCaptcha aparecer no DOM
            logger.info("Aguardando campo do hCaptcha aparecer...")
            page.wait_for_selector("textarea[name='h-captcha-response']", timeout=20000)

            # Injetar token resolvido no campo correto
            logger.info("Injetando token resolvido no campo do hCaptcha...")
            page.evaluate(f"""
                const el = document.querySelector("textarea[name='h-captcha-response']");
                if (el) {{
                    el.style.display = 'block';
                    el.value = "{token}";
                    el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                    el.dispatchEvent(new Event('change', {{ bubbles: true }}));
                }} else {{
                    throw new Error("Campo h-captcha-response não encontrado no DOM");
                }}
            """)
            logger.success("Token do hCaptcha injetado com sucesso.")

            # Preenche senha
            logger.info("Preenchendo senha...")
            page.get_by_role("textbox", name="Senha").fill(SENHA_LOGIN)

            # Enviar formulário
            page.get_by_role("button", name="Entrar").click()
            logger.info("Login realizado, aguardando redirecionamento...")

            time.sleep(10)  # ajuste conforme necessário

        except Exception as e:
            motivo = f"{type(e).__name__}: {e}"
            logger.error(f"Erro durante processo de login: {motivo}")
            traceback.print_exc()

        finally:
            context.close()
            browser.close()

    logger.info("Processo concluído.")


# === Execução ===
if __name__ == "__main__":
    processar_mte()
