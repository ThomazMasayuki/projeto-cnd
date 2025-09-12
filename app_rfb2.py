import re
import time
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import fitz  
from loguru import logger
from playwright.sync_api import sync_playwright
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "RFB"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
COL_STATUS = "STATUS"

OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

URL_RFB = "https://servicos.receitafederal.gov.br/servico/certidoes/#/home/cnpj"
REGEX_VALIDADE = r"VÁLIDA ATÉ:\s*(\d{2}/\d{2}/\d{4})"


# === Funções utilitárias ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)


def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""


def salvar_valor_na_planilha(cnpj: str, valor: str, caminho: Path, aba: str, coluna: str = COL_VALIDADE):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_coluna = colunas.get(coluna)

    if not idx_cnpj or not idx_coluna:
        logger.error(f"Coluna {coluna} não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_coluna - 1].value = valor
            break

    wb.save(caminho)
    wb.close()


# === Helpers Playwright ===
def clicar_nova_certidao(page_like):
    """page_like pode ser 'page' ou 'frame'. Tenta várias formas do botão '+ Nova Certidão'."""
    try:
        btn = page_like.get_by_role("button", name=re.compile(r"^\s*\+\s*Nova Certidão\s*$", re.I))
        if btn.count() and btn.first.is_visible():
            btn.first.click()
            return
    except Exception:
        pass
    try:
        page_like.locator("button:has-text('+ Nova Certidão')").first.click()
        return
    except Exception:
        pass
    try:
        page_like.locator("button.br-button.primary.btn-acao").first.click()
        return
    except Exception:
        pass
    page_like.locator("xpath=//button[contains(.,'+ Nova Certidão')]").first.click()


def preencher_cnpj_robusto(page, cnpj_limpo: str, timeout_ms: int = 20000) -> None:
    """
    Encontra e preenche o campo de CNPJ por múltiplas estratégias.
    Se não encontrar, levanta RuntimeError com explicação.
    """
    # 1) Seletores simples
    seletores = [
        "input[name='niContribuinte']",
        "input[placeholder='Informe o CNPJ']",
        "input[type='text'][maxlength='18']",
        "input[maxlength='18']",
    ]
    for css in seletores:
        loc = page.locator(css).first
        try:
            loc.wait_for(state="visible", timeout=timeout_ms)
            loc.scroll_into_view_if_needed()
            loc.click()
            loc.fill("")                # limpa
            loc.type(cnpj_limpo, delay=300)
            logger.info(f"[RFB] CNPJ preenchido via seletor: {css}")
            return
        except Exception:
            pass

    # 2) XPaths baseados no rótulo "CNPJ"
    xpaths = [
        "xpath=//label[normalize-space()='CNPJ']/following::input[1]",
        "xpath=//*[contains(normalize-space(.),'CNPJ')]/following::input[1]",
        "xpath=(//input)[1][ancestor::*[contains(.,'CNPJ')]]",
    ]
    for xp in xpaths:
        loc = page.locator(xp).first
        try:
            loc.wait_for(state="visible", timeout=timeout_ms)
            bb = loc.bounding_box()
            if bb:
                page.mouse.move(bb["x"] + bb["width"]/2, bb["y"] + bb["height"]/2)
                page.mouse.click(bb["x"] + bb["width"]/2, bb["y"] + bb["height"]/2)
            else:
                loc.click()
            loc.fill("")
            loc.type(cnpj_limpo, delay=300)
            logger.info(f"[RFB] CNPJ preenchido via XPath: {xp}")
            return
        except Exception:
            pass

    # 3) Varredura via JS + clique por bounding box (último recurso)
    logger.info("[RFB] Tentando localizar campo via varredura JS...")
    res = page.evaluate(
        """
        (cnpj) => {
          // Candidatos por nome/placeholder/comprimento
          const candidates = [];
          const byName = document.querySelector("input[name='niContribuinte']");
          if (byName) candidates.push(byName);
          for (const el of Array.from(document.querySelectorAll("input,textarea"))) {
            const ph = (el.getAttribute('placeholder') || '').toLowerCase();
            const nm = (el.getAttribute('name') || '').toLowerCase();
            const ml = el.maxLength || el.getAttribute('maxlength');
            const isTextish = (el.type || '').toLowerCase() in {text:1, tel:1, search:1, '' :1} || !el.type;
            if (!isTextish) continue;
            if (ph.includes('cnpj') || nm.includes('nicontribuinte') || ml == 18) {
              candidates.push(el);
            }
          }
          for (const el of candidates) {
            try {
              el.scrollIntoView({block:'center', inline:'center'});
              el.focus();
              el.value = '';
              el.dispatchEvent(new Event('input', {bubbles:true}));
              el.value = cnpj;
              el.dispatchEvent(new Event('input', {bubbles:true}));
              el.dispatchEvent(new Event('change', {bubbles:true}));
              const r = el.getBoundingClientRect();
              return { ok:true, x: r.left + r.width/2, y: r.top + r.height/2 };
            } catch(e) {}
          }
          return { ok:false };
        }
        """,
        cnpj_limpo,
    )
    if res and res.get("ok"):
        # clica no centro do campo para garantir foco/validações Angular
        page.mouse.move(res["x"], res["y"])
        page.mouse.click(res["x"], res["y"])
        logger.info("[RFB] CNPJ preenchido via varredura JS + clique.")
        return

    raise RuntimeError("Campo CNPJ não encontrado (mesmo com varredura JS).")


# === Função principal ===
def processar_rfb():
    logger.add("execucao_rfb.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")

    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ]].dropna(subset=[COL_CNPJ])
    cnpjs = df[COL_CNPJ].drop_duplicates().tolist()

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True, user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/114.0.0.0 Safari/537.36",
                       viewport={"width": 1920, "height": 1080})
        page = context.new_page()

        page.goto(URL_RFB, timeout=60000)
        # Aguarda o container principal da aplicação carregar
        page.wait_for_selector("input[placeholder='Informe o CNPJ']", timeout=60000)
        # ou, se esse input demorar para aparecer:
        # page.wait_for_selector("text=CNPJ", timeout=60000)

        logger.info("Página carregada e campo de CNPJ disponível.")

        for cnpj in cnpjs:
            cnpj_limpo = normalizar_cnpj(cnpj)
            logger.info(f"Processando {cnpj_limpo}...")

            try:
                # localizar e preencher (robusto)
                time.sleep(5)
                preencher_cnpj_robusto(page, cnpj_limpo)
                time.sleep(0.5)

                # clicar + Nova Certidão
                clicar_nova_certidao(page)

                # aguardar resultado
                time.sleep(5)

                # tratar mensagens de erro
                if page.locator("div.msg-resultado:has-text('Não foi possível concluir')").is_visible():
                    salvar_valor_na_planilha(cnpj_limpo, "ERRO BAIXAR", PLANILHA, ABA, COL_STATUS)
                    logger.warning(f"{cnpj_limpo} → ERRO BAIXAR")
                    continue

                # tratar confirmação de certidão existente
                if page.locator("text=Certidão Válida Encontrada").is_visible():
                    logger.info(f"{cnpj_limpo} → Certidão já existente. Reemitindo...")
                    clicar_nova_certidao(page)
                    time.sleep(5)

                # sucesso
                if page.locator("div.msg-resultado:has-text('A certidão foi emitida com sucesso')").is_visible():
                    logger.success(f"{cnpj_limpo} → Certidão emitida com sucesso.")
                    salvar_valor_na_planilha(cnpj_limpo, datetime.today().strftime("%d/%m/%Y"), PLANILHA, ABA)

                    # renomear downloads
                    downloads = context.downloads
                    for d in downloads:
                        path = d.path()
                        if not path:
                            continue
                        destino = OUTPUT_DIR / f"rfb_{cnpj_limpo}_{datetime.today().strftime('%Y%m%d')}.pdf"
                        Path(path).rename(destino)
                        logger.info(f"Salvo em {destino}")

                # preparar próxima iteração
                clicar_nova_certidao(page)

            except Exception as e:
                logger.error(f"{cnpj_limpo} → ERRO: {e}")
                traceback.print_exc()
                salvar_valor_na_planilha(cnpj_limpo, "ERRO SCRIPT", PLANILHA, ABA, COL_STATUS)

        context.close()
        browser.close()

    logger.info("Processo concluído.")

if __name__ == "__main__":
    processar_rfb()
