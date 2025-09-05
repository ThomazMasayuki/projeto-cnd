import re
import time
import traceback
import random
import tempfile
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional, Dict

import cv2
import numpy as np
import pandas as pd
import fitz  # PyMuPDF
from PIL import Image
import pytesseract

from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook

# =========================
# === Configurações ===
# =========================
PLANILHA = Path("base_certidoes.xlsx")
ABA = "CDT"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"

OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_CDT = "https://cndt-certidao.tst.jus.br/gerarCertidao.faces"
TIMEOUT = 40_000
REGEX_VALIDADE = r"VÁLIDA ATÉ:\s*(\d{2}/\d{2}/\d{4})"

# === Captcha ===
EXPECTED_LEN = 5  # ajustar para 5 ou 6 conforme o captcha do TST
CHARSET = "ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz0123456789"
TESS_LANG = "eng"
SAVE_DEBUG = False  # True salva variações em /tmp
MAX_PIPELINES = 6
CONF_MIN = 65  # confiança mínima média por leitura (0-100)
MAX_TENTATIVAS_CAPTCHA = 5

# === Selectors (ajustar conforme DOM real do TST) ===
SEL_CNPJ_INPUT = "input[name='formConsultaCertidao:numeroDocumento']"  # exemplo: ajustável
SEL_CAPTCHA_IMG = "img[id='formConsultaCertidao:captchaImg']"         # exemplo: ajustável
SEL_CAPTCHA_INPUT = "input[name='formConsultaCertidao:captcha']"      # exemplo: ajustável
SEL_EMITIR_BTN = "button[id='formConsultaCertidao:btnEmitir']"        # exemplo: ajustável
SEL_CAPTCHA_REFRESH = ""  # se existir (ex.: "a#refreshCaptcha")

# =========================
# === Utilitários ===
# =========================
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDÃO não encontrada.")
        wb.close()
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

def carregar_imagem_com_checagem(caminho_imagem: Path) -> np.ndarray:
    if not caminho_imagem.exists():
        raise FileNotFoundError(f"Imagem não encontrada: {caminho_imagem}")
    img = cv2.imread(str(caminho_imagem), cv2.IMREAD_COLOR)
    if img is None or img.size == 0:
        raise ValueError(f"Falha ao carregar imagem: {caminho_imagem}")
    return img

def _debug_save(img: np.ndarray, name: str):
    if not SAVE_DEBUG:
        return
    p = Path(tempfile.gettempdir()) / f"captcha_{int(time.time()*1000)}_{name}.png"
    cv2.imwrite(str(p), img)

def _to_gray(img: np.ndarray) -> np.ndarray:
    return cv2.cvtColor(img, cv2.COLOR_BGR2GRAY) if len(img.shape) == 3 else img

def _resize(img: np.ndarray, scale: float = 2.8) -> np.ndarray:
    return cv2.resize(img, None, fx=scale, fy=scale, interpolation=cv2.INTER_CUBIC)

def _deskew_light(gray: np.ndarray) -> np.ndarray:
    thr = cv2.threshold(gray, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)[1]
    coords = np.column_stack(np.where(thr > 0))
    angle = 0.0
    if len(coords) > 10:
        rect = cv2.minAreaRect(coords)
        angle = rect[-1]
        if angle < -45:
            angle = -(90 + angle)
        else:
            angle = -angle
    if abs(angle) < 0.5:
        return gray
    (h, w) = gray.shape[:2]
    M = cv2.getRotationMatrix2D((w//2, h//2), angle, 1.0)
    rotated = cv2.warpAffine(gray, M, (w, h), flags=cv2.INTER_CUBIC, borderMode=cv2.BORDER_REPLICATE)
    return rotated

def _remove_small_specks(bin_img: np.ndarray, min_area: int = 20) -> np.ndarray:
    nb_components, output, stats, _ = cv2.connectedComponentsWithStats(255 - bin_img, connectivity=8)
    sizes = stats[1:, -1]
    img2 = np.copy(bin_img)
    for i, size in enumerate(sizes, start=1):
        if size < min_area:
            img2[output == i] = 255
    return img2

def _remove_lines(bin_img: np.ndarray) -> np.ndarray:
    horiz = bin_img.copy()
    vert = bin_img.copy()
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (25, 1))
    v_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (1, 25))
    h_tophat = cv2.morphologyEx(255 - horiz, cv2.MORPH_TOPHAT, h_kernel)
    v_tophat = cv2.morphologyEx(255 - vert, cv2.MORPH_TOPHAT, v_kernel)
    clean = bin_img.copy()
    clean[h_tophat > 0] = 255
    clean[v_tophat > 0] = 255
    return clean

def _clahe(gray: np.ndarray) -> np.ndarray:
    clahe = cv2.createCLAHE(clipLimit=3.0, tileGridSize=(8,8))
    return clahe.apply(gray)

def _adaptive_binarize(gray: np.ndarray) -> np.ndarray:
    return cv2.adaptiveThreshold(gray, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                 cv2.THRESH_BINARY, 31, 10)

def _otsu_binarize(gray: np.ndarray, invert: bool = False) -> np.ndarray:
    flag = cv2.THRESH_BINARY_INV if invert else cv2.THRESH_BINARY
    thr = cv2.threshold(gray, 0, 255, flag + cv2.THRESH_OTSU)[1]
    return thr

def _morph_refine(bin_img: np.ndarray) -> np.ndarray:
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (2,2))
    bin_img = cv2.morphologyEx(bin_img, cv2.MORPH_CLOSE, kernel, iterations=1)
    bin_img = cv2.medianBlur(bin_img, 3)
    return bin_img

def _invert_if_needed(bin_img: np.ndarray) -> np.ndarray:
    black_pixels = np.sum(bin_img == 0)
    white_pixels = np.sum(bin_img == 255)
    if black_pixels < white_pixels:
        return 255 - bin_img
    return bin_img

def _tesseract_read(pil_img: Image.Image, psm: int) -> Tuple[str, float]:
    config = f'--oem 3 --psm {psm} -c tessedit_char_whitelist={CHARSET}'
    data = pytesseract.image_to_data(pil_img, lang=TESS_LANG, config=config, output_type=pytesseract.Output.DICT)
    texts, confs = [], []
    for t, c in zip(data.get('text', []), data.get('conf', [])):
        if t and t.strip() and str(c).isdigit():
            texts.append(t.strip())
            confs.append(int(c))
    raw = "".join(texts)
    conf = float(np.mean(confs)) if confs else -1.0
    raw = re.sub(r'[^A-Za-z0-9]', '', raw)
    return raw, conf

def _pipelines(gray: np.ndarray) -> List[np.ndarray]:
    vars = []

    g1 = _clahe(gray)
    v1 = _otsu_binarize(g1, invert=False)
    v1 = _remove_small_specks(v1, min_area=40)
    v1 = _remove_lines(v1)
    v1 = _morph_refine(v1)
    v1 = _invert_if_needed(v1)
    vars.append(v1)

    g2 = _clahe(gray)
    v2 = _otsu_binarize(g2, invert=True)
    v2 = _remove_small_specks(v2, min_area=40)
    v2 = _morph_refine(v2)
    v2 = _invert_if_needed(v2)
    vars.append(v2)

    a1 = _adaptive_binarize(gray)
    a1 = _remove_small_specks(a1, min_area=50)
    a1 = _morph_refine(a1)
    a1 = _invert_if_needed(a1)
    vars.append(a1)

    o1 = _otsu_binarize(gray, invert=False)
    o1 = _remove_lines(o1)
    o1 = _morph_refine(o1)
    o1 = _invert_if_needed(o1)
    vars.append(o1)

    m1 = cv2.medianBlur(gray, 5)
    a2 = _adaptive_binarize(m1)
    a2 = _remove_small_specks(a2, min_area=60)
    a2 = _morph_refine(a2)
    a2 = _invert_if_needed(a2)
    vars.append(a2)

    b1 = cv2.bilateralFilter(gray, 9, 75, 75)
    o2 = _otsu_binarize(b1, invert=False)
    o2 = _remove_small_specks(o2, min_area=40)
    o2 = _morph_refine(o2)
    o2 = _invert_if_needed(o2)
    vars.append(o2)

    return vars[:MAX_PIPELINES]

def solver_captcha_ensemble(caminho_imagem: Path,
                            expected_len: int = EXPECTED_LEN,
                            try_psms: List[int] = [7,8,13]) -> Dict:
    img = carregar_imagem_com_checagem(caminho_imagem)
    gray = _to_gray(img)
    gray = _deskew_light(gray)
    gray = _resize(gray, scale=2.8)

    candidates = []
    for idx, var in enumerate(_pipelines(gray)):
        _debug_save(var, f"var_{idx}.png")
        pil = Image.fromarray(var)
        for psm in try_psms:
            text, conf = _tesseract_read(pil, psm=psm)
            if text:
                candidates.append((text, conf, psm, idx))

    def score_item(t, c):
        length_penalty = 0.0 if expected_len <= 0 else -abs(len(t) - expected_len) * 2.0
        return c + length_penalty

    agg: Dict[str, float] = {}
    for t, c, _, _ in candidates:
        s = score_item(t, c)
        agg[t] = agg.get(t, 0.0) + max(s, -50)

    best_text = ""
    if agg:
        best_text = max(agg.items(), key=lambda kv: kv[1])[0]

    if expected_len > 0 and best_text and len(best_text) != expected_len:
        ok_alts = [(t, s) for t, s in agg.items() if len(t) == expected_len]
        if ok_alts:
            best_text = max(ok_alts, key=lambda kv: kv[1])[0]

    best_conf = -1.0
    for t, c, psm, idx in candidates:
        if t == best_text:
            best_conf = max(best_conf, c)

    return {"best_text": best_text, "best_conf": best_conf, "candidates": candidates}

def possivel_operacao_aritmetica(texto: str) -> Optional[int]:
    m = re.fullmatch(r'\s*(\d+)\s*([+\-xX*])\s*(\d+)\s*=?\s*', texto or "")
    if not m:
        return None
    a, op, b = m.groups()
    a = int(a); b = int(b)
    if op in ['x', 'X', '*']:
        return a * b
    elif op == '+':
        return a + b
    else:
        return a - b

def resolver_captcha(page, captcha_selector: str, refresh_selector: str = "") -> str:
    """
    Captura a imagem do captcha e roda o ensemble por até MAX_TENTATIVAS_CAPTCHA.
    Se houver 'refresh_selector', clica para trocar o captcha; caso contrário, recarrega a página.
    Retorna o texto do captcha (ou string vazia se falhar).
    """
    best_global = {"best_text": "", "best_conf": -1.0}
    temp_dir = Path(tempfile.gettempdir())

    for tentativa in range(1, MAX_TENTATIVAS_CAPTCHA + 1):
        el = page.locator(captcha_selector).first
        el.wait_for(state="visible", timeout=8_000)
        img_path = temp_dir / f"captcha_try{tentativa}_{int(time.time()*1000)}.png"
        el.screenshot(path=str(img_path))

        result = solver_captcha_ensemble(img_path, expected_len=EXPECTED_LEN)
        logger.info(f"[Captcha tentativa {tentativa}] Lido='{result['best_text']}' (conf={result['best_conf']:.1f})")

        if result["best_conf"] > best_global["best_conf"]:
            best_global = result

        # Operação aritmética? (caso o site use)
        op_val = possivel_operacao_aritmetica(result["best_text"])
        if op_val is not None:
            return str(op_val)

        # Critério de aceitação
        if result["best_text"] and (EXPECTED_LEN <= 0 or len(result["best_text"]) == EXPECTED_LEN) and (result["best_conf"] >= CONF_MIN):
            return result["best_text"]

        # Refresh
        try:
            if refresh_selector:
                page.locator(refresh_selector).click()
            else:
                page.reload()
        except Exception:
            page.reload()
        time.sleep(0.8 + random.random() * 0.6)

    return best_global["best_text"] or ""

# =========================
# === Fluxo principal ===
# =========================
def processar_cdt():
    logger.add("execucaocdt.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")

    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ, COL_VALIDADE]].dropna(subset=[COL_CNPJ])
    cnpjs = df[COL_CNPJ].drop_duplicates().tolist()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        for cnpj in cnpjs:
            cnpj_limpo = normalizar_cnpj(cnpj)

            try:
                logger.info(f"Consultando CNPJ: {cnpj_limpo}")
                page.goto(URL_CDT, timeout=TIMEOUT)
                page.wait_for_load_state("domcontentloaded", timeout=10_000)

                # Preenche o CNPJ (ajustar selector se necessário)
                page.locator(SEL_CNPJ_INPUT).fill(cnpj_limpo)

                # Aguarda captcha
                page.wait_for_selector(SEL_CAPTCHA_IMG, timeout=10_000)
                time.sleep(1.0)

                # Resolve captcha (OCR robusto)
                texto_captcha = resolver_captcha(page, SEL_CAPTCHA_IMG, SEL_CAPTCHA_REFRESH)
                if not texto_captcha:
                    logger.warning(f"{cnpj_limpo} → Não foi possível ler o captcha com confiança.")
                page.locator(SEL_CAPTCHA_INPUT).fill(texto_captcha)

                time.sleep(0.5)

                # Emite certidão (espera nova página se o portal abrir em outra aba)
                with context.expect_page() as nova_aba_evento:
                    page.locator(SEL_EMITIR_BTN).click()

                nova_aba = nova_aba_evento.value
                nova_aba.wait_for_load_state("networkidle", timeout=15_000)

                # Salva PDF da nova aba
                temp_path = OUTPUT_DIR / f"temp_{cnpj_limpo}.pdf"
                try:
                    # Observação: page.pdf depende do navegador/headless. Se houver erro, capturar e salvar screenshot como fallback.
                    nova_aba.pdf(path=str(temp_path), format="A4")
                except Exception as e_pdf:
                    logger.warning(f"{cnpj_limpo} → Falha em gerar PDF direto ({e_pdf}). Salvando screenshot full-page como fallback.")
                    screenshot_path = OUTPUT_DIR / f"screenshot_{cnpj_limpo}.png"
                    nova_aba.screenshot(path=str(screenshot_path), full_page=True)
                    # Tenta converter via print-to-pdf do próprio site, se existir, ou mantém o PNG.

                # Extrai validade
                validade = ""
                if temp_path.exists():
                    validade = extrair_validade_pdf(temp_path)

                if validade:
                    salvar_valor_na_planilha(cnpj_limpo, validade, PLANILHA, ABA)
                    validade_formatada = datetime.strptime(validade, "%d/%m/%Y").strftime("%Y%m%d")
                    nome_arquivo = f"cdt_{cnpj_limpo}_{validade_formatada}.pdf"
                    destino_pdf = OUTPUT_DIR / nome_arquivo
                    if temp_path.exists():
                        temp_path.rename(destino_pdf)
                    logger.success(f"{cnpj_limpo} → Sucesso: validade {validade}")
                else:
                    # Guarda o artefato para inspeção
                    if temp_path.exists():
                        destino_pdf = OUTPUT_DIR / f"erro_{cnpj_limpo}.pdf"
                        temp_path.rename(destino_pdf)
                    logger.warning(f"{cnpj_limpo} → Não foi possível extrair validade (verificar arquivo gerado).")

                nova_aba.close()
                page.bring_to_front()

            except Exception as e:
                motivo = f"{type(e).__name__}: {e}"
                logger.error(f"{cnpj_limpo} → ERRO: {motivo}")
                traceback.print_exc()

        context.close()
        browser.close()

    logger.info("Processo concluído.")

# === Execução ===
if __name__ == "__main__":
    processar_cdt()