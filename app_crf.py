import os
import re
import time
import base64
import traceback
from datetime import datetime
from pathlib import Path
import requests

import pandas as pd
from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook

# =====================
# Configurações
# =====================
PLANILHA = Path("base_certidoes.xlsx")
ABA = "CRF"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
COL_STATUS = "STATUS"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_CRF = "https://consulta-crf.caixa.gov.br/consultacrf/pages/consultaEmpregador.jsf"
TIMEOUT = 40_000
REGEX_VALIDADE_FINAL = r"Validade:\s*\d{2}/\d{2}/\d{4}\s*a\s*(\d{2}/\d{2}/\d{4})"

API_KEY_2CAPTCHA = os.getenv("API_KEY_2CAPTCHA", "30924a201f06a3b554d7479c487fee8e")
POLLING_2CAPTCHA_SEG = 5
MAX_POLLS_2CAPTCHA = 50 
MAX_TENTATIVAS_CNPJ = 6
HEADLESS = False        

# =====================
# Utilitários planilha (openpyxl)
# =====================

def _abrir_ws(caminho: Path, aba: str):
    wb = load_workbook(caminho)
    if aba not in wb.sheetnames:
        raise ValueError(f"Aba '{aba}' não encontrada em {caminho}.")
    ws = wb[aba]
    return wb, ws


def _mapear_cabecalhos(ws):
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    return colunas


def _garantir_coluna_status(ws, colunas):
    if COL_STATUS not in colunas:
        # cria na próxima coluna
        nova_idx = ws.max_column + 1
        ws.cell(row=1, column=nova_idx, value=COL_STATUS)
        colunas[COL_STATUS] = nova_idx


def salvar_validade_status_na_planilha(cnpj: str, validade: str | None, status: str):
    wb, ws = _abrir_ws(PLANILHA, ABA)
    colunas = _mapear_cabecalhos(ws)

    # Garante coluna VALIDADE e STATUS existirem
    if COL_CNPJ not in colunas:
        raise ValueError(f"Coluna '{COL_CNPJ}' não encontrada na aba {ABA}.")
    if COL_VALIDADE not in colunas:
        # cria VALIDADE se não existir
        nova_idx = ws.max_column + 1
        ws.cell(row=1, column=nova_idx, value=COL_VALIDADE)
        colunas[COL_VALIDADE] = nova_idx

    _garantir_coluna_status(ws, colunas)

    idx_cnpj = colunas[COL_CNPJ]
    idx_validade = colunas[COL_VALIDADE]
    idx_status = colunas[COL_STATUS]

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if re.sub(r"\D", "", str(val)).zfill(14) == re.sub(r"\D", "", str(cnpj)).zfill(14):
            if validade:
                row[idx_validade - 1].value = validade
            row[idx_status - 1].value = status
            break

    wb.save(PLANILHA)
    wb.close()

# =====================
# 2Captcha – image captcha
# =====================

def resolver_captcha_2captcha(caminho_imagem: Path, api_key: str) -> str:
    with open(caminho_imagem, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    payload = {"key": api_key, "method": "base64", "body": b64, "json": 1}
    resp = requests.post("http://2captcha.com/in.php", data=payload, timeout=40)
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") != 1:
        raise RuntimeError(f"Falha ao enfileirar captcha: {data}")

    captcha_id = data["request"]

    for _ in range(MAX_POLLS_2CAPTCHA):
        time.sleep(POLLING_2CAPTCHA_SEG)
        res = requests.get(
            "http://2captcha.com/res.php",
            params={"key": api_key, "action": "get", "id": captcha_id, "json": 1},
            timeout=40,
        )
        res.raise_for_status()
        ans = res.json()
        if ans.get("status") == 1:
            return ans["request"].strip()
        if ans.get("request") != "CAPCHA_NOT_READY":
            raise RuntimeError(f"Erro 2Captcha: {ans}")

    raise TimeoutError("Timeout aguardando solução do 2Captcha")

# =====================
# Coleta de validade (página HTML final)
# =====================

def extrair_validade_do_html(page_html: str) -> str | None:
    m = re.search(REGEX_VALIDADE_FINAL, page_html, flags=re.IGNORECASE)
    if m:
        return m.group(1)  # data final
    return None

# =====================
# Fluxo principal (FGTS/CRF)
# =====================

def processar_crf():
    logger.add("execucaocrf.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")

    if not API_KEY_2CAPTCHA or API_KEY_2CAPTCHA == "COLOQUE_SUA_CHAVE_AQUI":
        logger.error("Defina API_KEY_2CAPTCHA (env API_KEY_2CAPTCHA) antes de executar.")
        return

    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ]].dropna(subset=[COL_CNPJ])
    cnpjs = df[COL_CNPJ].drop_duplicates().tolist()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS, args=["--start-maximized"])
        context = browser.new_context(accept_downloads=True, no_viewport=True)
        page = context.new_page()

        for cnpj in cnpjs:
            cnpj_limpo = re.sub(r"\D", "", str(cnpj)).zfill(14)
            tentativas = 0
            status_final = "FALHA"
            validade_final = None

            while tentativas < MAX_TENTATIVAS_CNPJ:
                tentativas += 1
                try:
                    logger.info(f"Consultando CRF (FGTS) – CNPJ {cnpj_limpo} [tentativa {tentativas}/{MAX_TENTATIVAS_CNPJ}]")
                    page.goto(URL_CRF, timeout=TIMEOUT)
                    page.wait_for_load_state("domcontentloaded", timeout=1500)

                    # --- Seleciona CNPJ e preenche inscrição ---
                    radio_ok = False
                    try:
                        page.get_by_label(re.compile(r"\bCNPJ\b", re.I)).check()
                        radio_ok = True
                    except Exception:
                        try:
                            lbl = page.locator("label:has-text('CNPJ')").first
                            lbl.wait_for(state="visible", timeout=8000)
                            lbl.click()
                            radio_ok = True
                        except Exception:
                            try:
                                page.locator("xpath=//label[contains(normalize-space(),'CNPJ')]").first.click()
                                radio_ok = True
                            except Exception:
                                radio_ok = False

                    if not radio_ok:
                        try:
                            page.evaluate("""
                                (() => {
                                    const lbl = [...document.querySelectorAll('label')]
                                      .find(l => /\\bCNPJ\\b/i.test(l.textContent || ''));
                                    if (!lbl) return false;
                                    const forId = lbl.getAttribute('for');
                                    let input = null;
                                    if (forId) input = document.getElementById(forId);
                                    if (!input) {
                                      input = lbl.previousElementSibling && lbl.previousElementSibling.type === 'radio'
                                        ? lbl.previousElementSibling
                                        : (lbl.nextElementSibling && lbl.nextElementSibling.type === 'radio'
                                          ? lbl.nextElementSibling : null);
                                    }
                                    if (!input) return false;
                                    input.checked = true;
                                    input.dispatchEvent(new Event('change', {bubbles: true}));
                                    input.dispatchEvent(new Event('input', {bubbles: true}));
                                    return true;
                                })();
                            """)
                        except Exception:
                            pass

                    campo = page.locator("#mainForm\\:txtInscricao1")
                    campo.wait_for(state="visible", timeout=8000)
                    page.wait_for_timeout(400)
                    try:
                        campo.click()
                        campo.fill("")
                        campo.type(cnpj_limpo, delay=30)
                        campo.press("Tab")
                    except Exception as e:
                        logger.debug(f"Falha ao digitar no campo Inscrição: {e}")

                    try:
                        valor = campo.input_value(timeout=2000)
                    except Exception:
                        valor = ""
                    if re.sub(r"\\D", "", valor) != cnpj_limpo:
                        try:
                            page.evaluate(
                                """(sel, val) => {
                                    const el = document.querySelector(sel);
                                    if (!el) return;
                                    el.focus();
                                    el.value = val;
                                    el.dispatchEvent(new Event('input', { bubbles: true }));
                                    el.dispatchEvent(new Event('change', { bubbles: true }));
                                    el.blur();
                                }""",
                                "#mainForm\\:txtInscricao1", cnpj_limpo
                            )
                            logger.info("Valor do campo Inscrição forçado via JS.")
                        except Exception as e:
                            logger.warning(f"Fallback JS para Inscrição falhou: {e}")

                    # --- Captura e resolve o captcha (2Captcha image) ---
                    time.sleep(1.5)
                    captcha_img = page.locator("img[alt*='captcha' i], img[src*='captcha']").first
                    captcha_path = OUTPUT_DIR / f"captcha_{cnpj_limpo}.png"
                    captcha_img.screenshot(path=str(captcha_path))

                    texto_captcha = resolver_captcha_2captcha(captcha_path, API_KEY_2CAPTCHA)
                    logger.info(f"2Captcha → '{texto_captcha}'")

                    # --- PREENCHE O CAPTCHA (campo id 'mainForm:txtCaptcha') ---
                    sel_cap = "#mainForm\\:txtCaptcha"
                    cap = page.locator(sel_cap)
                    cap.wait_for(state="visible", timeout=8000)
                    page.wait_for_timeout(150)
                    try:
                        cap.click()
                        cap.fill("")
                        cap.type(texto_captcha, delay=20)
                        cap.press("Tab")
                    except Exception as e:
                        logger.debug(f"Falha ao digitar no captcha: {e}")
                    try:
                        val_cap = cap.input_value(timeout=1000)
                    except Exception:
                        val_cap = ""
                    esperado = re.sub(r"\\W", "", texto_captcha or "").strip()
                    recebido = re.sub(r"\\W", "", val_cap or "").strip()
                    if recebido != esperado and esperado:
                        try:
                            page.evaluate(
                                """(sel, val) => {
                                    const el = document.querySelector(sel);
                                    if (!el) return;
                                    el.focus();
                                    el.value = val;
                                    el.dispatchEvent(new Event('input', { bubbles: true }));
                                    el.dispatchEvent(new Event('change', { bubbles: true }));
                                    el.blur();
                                }""",
                                sel_cap, esperado
                            )
                            logger.info("Captcha setado via JS (fallback).")
                        except Exception as e:
                            logger.warning(f"Fallback JS no captcha falhou: {e}")

                    # --- Consultar ---
                    page.get_by_role("button", name=re.compile("Consultar", re.I)).click()
                    page.wait_for_load_state("networkidle", timeout=20000)

                    # Verifica se apareceu o link do certificado
                    link_cert = page.get_by_role("link", name=re.compile("Certificado de Regularidade do FGTS - CRF", re.I))
                    if not link_cert.first.is_visible(timeout=3000):
                        screenshot_err = OUTPUT_DIR / f"crf_{cnpj_limpo}_erro_consulta.png"
                        page.screenshot(path=str(screenshot_err), full_page=True)
                        logger.warning("Consulta não retornou link do certificado; tentando novamente…")
                        try:
                            captcha_img.click()
                        except Exception:
                            pass
                        time.sleep(1.2)
                        continue

                    # Segue para o certificado
                    link_cert.first.click()
                    page.wait_for_load_state("networkidle", timeout=15000)

                    try:
                        page.get_by_role("button", name=re.compile("Visualizar", re.I)).click()
                    except Exception:
                        page.locator("#mainForm\\:btnVisualizar").click()

                    # Pode abrir nova aba ou ficar na mesma
                    temp_img = OUTPUT_DIR / f"crf_{cnpj_limpo}_certidao.png"
                    temp_pdf = OUTPUT_DIR / f"crf_{cnpj_limpo}_certidao.pdf"
                    try:
                        with context.expect_page(timeout=8000) as nova:
                            pass
                    except PWTimeout:
                        cert_page = page
                    else:
                        cert_page = nova.value
                        cert_page.wait_for_load_state("networkidle", timeout=1500)

                    # Evidências
                    try:
                        cert_page.screenshot(path=str(temp_img), full_page=True)
                    except Exception as e:
                        logger.debug(f"Falha ao tirar screenshot: {e}")
                    try:
                        cert_page.pdf(path=str(temp_pdf), format="A4")
                    except Exception:
                        pass

                    # Validade (HTML)
                    html = cert_page.content()
                    validade = extrair_validade_do_html(html)
                    if validade:
                        validade_final = validade

                    status_final = "OK"
                    logger.success(f"CNPJ {cnpj_limpo} → Sucesso (status OK)")
                    break

                except Exception as e:
                    motivo = f"{type(e).__name__}: {e}"
                    logger.error(f"{cnpj_limpo} → ERRO: {motivo}")
                    traceback.print_exc()
                    status_final = "ERRO"

            # Atualiza planilha
            salvar_validade_status_na_planilha(cnpj_limpo, validade_final, status_final)

        context.close()
        browser.close()

    logger.info("Processo concluído (CRF/FGTS).")


if __name__ == "__main__":
    processar_crf()
