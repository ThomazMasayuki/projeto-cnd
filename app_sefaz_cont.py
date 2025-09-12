import re
import time
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import fitz  # PyMuPDF
from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "SEFAZ CONT"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_SEFAZ_CONT = "https://sistemas.sefaz.am.gov.br/GAE/mnt/dividaAtiva/certidaoNegativa/emitirCertidaoNegativaNaoContPortal.do"
TIMEOUT = 40_000
REGEX_VALIDADE = r"Válida até:\s*(\d{2}/\d{2}/\d{4})"

def limpar_cnpj(doc: str) -> str:
    return re.sub(r"D", "", str(doc))

def extrair_validade_pdf(caminho_pdf:Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(doc: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if limpar_cnpj(val) == limpar_cnpj(doc):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()
    
def processar_sefaz_contribuinte():
    logger.add("execucao.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")
    
    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ, COL_VALIDADE]].dropna(subset=[COL_CNPJ])
    documentos = df[COL_CNPJ].drop_duplicates().to_list()
    
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()
        
        for doc_bruto in documentos:
            doc = limpar_cnpj(doc_bruto)
            
            if len(doc) == 11:
                tipo = "CPF"
                
            elif len(doc) == 14:
                tipo = "CNPJ"
                
            else: 
                logger.warning(f"{doc_bruto} -> Documento inválido (não é um CPF nem CNPJ). Pulando.")
                continue
            
            try: 
                logger.info(f"Consultando {tipo}: {doc}")
                page.goto(URL_SEFAZ_CONT, timeout=TIMEOUT)
                page.get_by_label("").fill(doc)
