import re
import time
import base64
import traceback
from datetime import datetime
from pathlib import Path
import requests

import pandas as pd
import fitz  
from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook
from uuid import uuid4

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "PMM"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_PMM = "https://semefatende.manaus.am.gov.br/servicoJanela.php?servico=1412"
TIMEOUT = 40_000
REGEX_VALIDADE = r"VÁLIDA ATÉ \s*(\d{2}/\d{2}/\d{4})"

# === Funções utilitárias ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

def resolver_captcha_2captcha(caminho_imagem: Path, api_key: str) -> str:
    with open(caminho_imagem, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")
    payload = {"key": api_key, "method": "base64", "body": b64, "json": 1}
    r = requests.post("http://2captcha.com/in.php", data=payload, timeout=40)
    r.raise_for_status()
    data = r.json()
    if data.get("status") != 1:
        raise RuntimeError(f"Falha ao enviar captcha: {data}")
    cap_id = data["request"]
    for _ in range(40):
        time.sleep(5)
        res = requests.get(
            "http://2captcha.com/res.php",
            params={"key": api_key, "action": "get", "id": cap_id, "json": 1},
            timeout=40,
        ).json()
        if res.get("status") == 1:
            return re.sub(r"\W", "", res["request"]).strip()
        if res.get("request") != "CAPCHA_NOT_READY":
            raise RuntimeError(f"Erro 2Captcha: {res}")
    raise TimeoutError("Timeout aguardando solução do 2Captcha")

def _log_frames(page):
    # opcional: ajuda a diagnosticar se há iframes
    try:
        frames = page.frames
        logger.debug(f"[diag] frames={len(frames)}")
        for i, fr in enumerate(frames):
            logger.debug(f"  #{i}: url={getattr(fr,'url',None)!r}, name={getattr(fr,'name',None)!r}")
    except Exception:
        pass

def _first_frame_with(page, css):
    # procura o seletor em todos os frames (raiz + iframes)
    try:
        if page.locator(css).count():
            return page
    except Exception:
        pass
    for fr in page.frames:
        try:
            if fr.locator(css).count():
                return fr
        except Exception:
            continue
    return None

def selecionar_radio_cnpj(page):
    """
    Seleciona o radio 'CNPJ' usando várias estratégias:
    id do input, label[for=...], role-accessible name e, por fim, click via JS.
    Retorna o frame (page ou iframe) onde o radio foi encontrado.
    """
    page.wait_for_load_state("domcontentloaded", timeout=15000)
    _log_frames(page)

    # id real visto no seu HTML: #VTIPOFILTRO3
    fr = _first_frame_with(page, "#VTIPOFILTRO3")
    if fr:
        radio = fr.locator("#VTIPOFILTRO3")
        radio.wait_for(state="attached", timeout=8000)

        # 1) check nativo
        try:
            radio.scroll_into_view_if_needed()
            radio.check()
            if radio.is_checked():
                return fr
        except Exception:
            pass

        # 2) clicar no label vinculado
        try:
            fr.locator("label[for='VTIPOFILTRO3']").scroll_into_view_if_needed()
            fr.locator("label[for='VTIPOFILTRO3']").click()
            if radio.is_checked():
                return fr
        except Exception:
            pass

        # 3) role + accessible name
        try:
            fr.get_by_role("radio", name=re.compile(r"\bCNPJ\b", re.I)).check()
            if radio.is_checked():
                return fr
        except Exception:
            pass

        # 4) JS direto (contorna overlay/handlers)
        try:
            fr.eval_on_selector("#VTIPOFILTRO3", "el => el.click()")
            if radio.is_checked():
                return fr
        except Exception:
            pass

    # fallback por texto do label (caso id mude)
    fr = _first_frame_with(page, "label:has-text('CNPJ')")
    if fr:
        fr.locator("label:has-text('CNPJ')").click()
        try:
            r2 = fr.get_by_role("radio", name=re.compile(r"\bCNPJ\b", re.I))
            if r2 and r2.is_checked():
                return fr
        except Exception:
            pass

    raise RuntimeError("Não foi possível selecionar o radio 'CNPJ'. Verifique se há iframe/overlay.")

def preencher_cnpj_no_campo(fr, cnpj_limpo):
    """
    Preenche o campo de CNPJ (id real visto: #VNRFILTRO) simulando digitação,
    e dá TAB para disparar onblur/validações do GeneXus.
    """
    campo = fr.locator("#vNRFILTRO")
    campo.wait_for(state="visible", timeout=10000)
    campo.scroll_into_view_if_needed()
    campo.click()
    campo.fill("")          # limpa
    campo.type(cnpj_limpo)  # digitação real lida melhor com máscaras
    campo.press("Tab")      # dispara onblur/validação
    time.sleep(0.3)

# seletores padrão (ajuste se necessário)
SEL_CAPTCHA_IMG = "img[src*='/Captcha/images/']"
SEL_CAPTCHA_INPUTS = [
    "#_cfield",                # id direto (prioritário)
    "input[name='_cfield']",   # redundância útil
]

def print_captcha(fr, out_path: Path = None) -> Path:
    el = fr.locator("img[src*='/Captcha/images/']").first
    el.wait_for(state="visible", timeout=15000)  # aumentamos o timeout aqui
    if not out_path:
        out_path = Path(f"./captcha_temp_{uuid4().hex[:6]}.png")
    out_path.parent.mkdir(parents=True, exist_ok=True)
    el.screenshot(path=str(out_path))
    return out_path

        
def preencher_captcha(fr, texto_captcha: str) -> None:
    for sel in SEL_CAPTCHA_INPUTS:
        try:
            loc = fr.locator(sel)
            if loc.count() and loc.is_visible():
                loc.scroll_into_view_if_needed()
                loc.fill("")  # limpa direto
                loc.fill(texto_captcha)
                loc.press("Tab")  # dispara validação GeneXus
                logger.info(f"[Captcha] Preenchido com '{texto_captcha}' via seletor '{sel}'")
                return
        except Exception as e:
            logger.warning(f"[Captcha] Falha ao tentar preencher com seletor {sel}: {e}")

    try:
        loc = fr.locator("label:has-text('Insira o código') ~ input").first
        loc.fill("")
        loc.fill(texto_captcha)
        loc.press("Tab")
    except Exception as e:
        raise RuntimeError(f"[Captcha] Não foi possível preencher com '{texto_captcha}': {e}")

# === Função principal ===
def processar_pmm():
    logger.add("execucaopmm.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")

    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ, COL_VALIDADE]].dropna(subset=[COL_CNPJ])
    cnpjs = df[COL_CNPJ].drop_duplicates().tolist()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(viewport={"width": 1920, "height": 1080})
        page = context.new_page()

        for cnpj in cnpjs:
            cnpj_limpo = normalizar_cnpj(cnpj)

            try:
                page.goto(URL_PMM, timeout=TIMEOUT)
                page.wait_for_load_state("domcontentloaded", timeout=15000)

                # Seleciona o radio CNPJ (retorna o frame correto)
                fr = selecionar_radio_cnpj(page)

                # Preenche o campo do CNPJ no mesmo frame
                preencher_cnpj_no_campo(fr, cnpj_limpo)

                # Localiza o frame que contém o captcha
                fr = _first_frame_with(page, "img[src*='/Captcha/images/']")
                if not fr:
                    raise RuntimeError("[Captcha] Não foi possível localizar o frame contendo a imagem do captcha.")

                # Captura e salva o captcha
                captcha_path = print_captcha(fr, OUTPUT_DIR / f"captcha_{cnpj_limpo}.png")

                # Resolve o captcha com 2Captcha
                texto_captcha = resolver_captcha_2captcha(captcha_path, API_KEY_2CAPTCHA)

                # Preenche o captcha no campo correto
                preencher_captcha(fr, texto_captcha)

                fr = _first_frame_with(page, "input[name='BTNCONSULTAR']")
                if not fr:
                    raise RuntimeError("[Botão] Não foi possível encontrar o frame com o botão 'Consultar'.")

                try:
                    fr.wait_for_selector("input[name='BTNCONSULTAR']", timeout=10000)

                    time.sleep(1)  
                    with context.expect_page(timeout=150000) as nova_pagina_evento:
                        fr.eval_on_selector("input[name='BTNCONSULTAR']", "el => el.click()")

                    time.sleep(5)  
                    nova_aba = nova_pagina_evento.value
                    nova_aba.wait_for_load_state("networkidle", timeout=150000)
                    logger.info(f"[Nova aba] Página carregada com sucesso: {nova_aba.url}")

                except PWTimeout:
                    raise RuntimeError("[Erro] A nova aba não foi aberta após clicar em 'Consultar' dentro de 30 segundos.")


                # Exporta o PDF e extrai validade
                # Verifica se a certidão não foi emitida por motivo de débito/restrição
                try:
                    alerta = nova_aba.locator("div.alert.alert-warning").text_content(timeout=5000)
                    if alerta and "não foi possível emitir a certidão" in alerta.lower():
                        salvar_valor_na_planilha(cnpj_limpo, "COM DÉBITO", PLANILHA, ABA)
                        logger.warning(f"{cnpj_limpo} → Certidão com débito detectada.")

                        # Salva a tela como evidência em PDF
                        temp_path = OUTPUT_DIR / f"pmm_{cnpj_limpo}_com_debito.pdf"
                        nova_aba.pdf(path=str(temp_path), format="A4")

                        nova_aba.close()
                        continue  # pula para o próximo CNPJ
                except Exception as e:
                    logger.debug(f"[Alerta] Nenhum alerta de débito encontrado: {e}")


                validade = extrair_validade_pdf(temp_path)
                if validade:
                    salvar_valor_na_planilha(cnpj_limpo, validade, PLANILHA, ABA)
                    validade_formatada = datetime.strptime(validade, "%d/%m/%Y").strftime("%Y%m%d")
                    nome_arquivo = f"pmm_{cnpj_limpo}_{validade_formatada}.pdf"
                    destino_pdf = OUTPUT_DIR / nome_arquivo
                    temp_path.rename(destino_pdf)
                    logger.success(f"{cnpj_limpo} → Sucesso: validade {validade}")
                else:
                    destino_pdf = OUTPUT_DIR / f"erro_{cnpj_limpo}.pdf"
                    temp_path.rename(destino_pdf)
                    logger.warning(f"{cnpj_limpo} → Não foi possível extrair validade.")

                nova_aba.close()

            except Exception as e:
                motivo = f"{type(e).__name__}: {e}"
                logger.error(f"{cnpj_limpo} → ERRO: {motivo}")
                traceback.print_exc()

        context.close()
        browser.close()

    logger.info("Processo concluído.")


if __name__ == "__main__":
    processar_pmm()
