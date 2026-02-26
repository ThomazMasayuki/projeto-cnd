from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
import time
import requests
import re
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import fitz  # PyMuPDF
from loguru import logger
from openpyxl import load_workbook

# === Configurações (TJAM Falência) ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "FALÊNCIA"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

URL_SITE = 'https://consultasaj.tjam.jus.br/sco/abrirCadastro.do'
REGEX_VALIDADE = r"VÁLIDA ATÉ:\s*(\d{2}/\d{2}/\d{4})"

# === Config Webmail / Roundcube ===
OUTPUT_EMAIL_DIR = Path("certidoes_email")
OUTPUT_EMAIL_DIR.mkdir(parents=True, exist_ok=True)

ASSUNTO_CERTIDAO = "Pedido de Certidão disponível para Download"
RE_LINK_TJAM = re.compile(r"^https://consultasaj\.tjam\.jus\.br", re.I)

# === Utilitários ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

# === 2Captcha ===
def solicitar_captcha(api_key, sitekey, url):
    payload = {'key': api_key, 'method': 'userrecaptcha', 'googlekey': sitekey, 'pageurl': url, 'json': 1}
    resposta = requests.post('http://2captcha.com/in.php', data=payload)
    return resposta.json().get('request')

def obter_resultado(api_key, captcha_id, tentativas=30, intervalo=7):
    for tentativa in range(tentativas):
        time.sleep(intervalo)
        payload = {'key': api_key, 'action': 'get', 'id': captcha_id, 'json': 0}
        resposta = requests.get('http://2captcha.com/res.php', params=payload)
        print(f"[DEBUG] Tentativa {tentativa+1} - Retorno da API: {resposta.text}")
        if 'OK|' in resposta.text:
            return resposta.text.split('|')[1]
        elif 'CAPCHA_NOT_READY' in resposta.text:
            continue
        else:
            raise ValueError(f"[2Captcha] Resposta inesperada: {resposta.text}")
    raise TimeoutError("Captcha não resolvido após várias tentativas.")

# === TJAM: preenchimento e envio ===
def automatizar_com_token(token_resolvido, cnpj: str, razao_social: str, context):
    page = context.new_page()
    page.goto(URL_SITE)
    page.wait_for_load_state('networkidle')

    page.wait_for_selector("select[name='entity.cdComarca']")
    page.select_option("select[name='entity.cdComarca']", value="1")   # Manaus
    page.select_option("select[name='entity.cdModelo']", value="31")   # Falência e Recuperação
    page.check("input[type='radio'][value='J']")                        # Pessoa Jurídica

    page.get_by_label("Razão Social", exact=False).fill(razao_social)
    page.get_by_label("CNPJ", exact=False).fill(cnpj)
    page.fill("input[name='entity.solicitante.deEmail']", EMAIL_PADRAO)

    # injeta token do recaptcha
    page.evaluate(f'''
        document.getElementById("g-recaptcha-response").style.display = 'block';
        document.getElementById("g-recaptcha-response").innerHTML = "{token_resolvido}";
    ''')

    page.check("input[type='checkbox'][value='true']")
    page.click("input[name='pbEnviar']")

    time.sleep(4)
    filename = f"{normalizar_cnpj(cnpj)}.png"
    page.screenshot(path=str(OUTPUT_DIR / filename), full_page=True)

    # volta para novo cadastro
    try:
        page.click("input[name='pbNovo']")
    except Exception:
        pass
    page.close()

# === Roundcube: baixar as certidões do dia (com print/pdf da página do TJAM) ===
def baixar_certidoes_email(context):
    page = context.new_page()
    logger.info("[Webmail] Acessando login...")
    page.goto(WEBMAIL_URL, timeout=60000)

    # login
    page.get_by_role("textbox", name="Endereço de e-mail").fill(EMAIL_USER)
    page.get_by_role("textbox", name="Senha").fill(EMAIL_PASS)
    page.get_by_role("button", name="Login").click()
    page.wait_for_load_state("networkidle", timeout=30000)
    logger.info("[Webmail] Login realizado.")

    # botão "Acessar" (open my active mail client)
    try:
        page.get_by_role("button", name=re.compile("Open my active mail client", re.I)).click()
    except Exception:
        logger.warning("[Webmail] Botão de acesso à caixa não encontrado (seguindo).")

    # garante que está na caixa de entrada
    try:
        page.get_by_role("link", name="Caixa de entrada").click()
    except Exception:
        pass
    time.sleep(2)

    hoje_label = "Hoje"  # Roundcube mostra "Hoje HH:MM"
    processed_ids = set()

    def buscar_rows_do_dia():
        # tr.unread = não lidos (negrito). Se não houver, pega lidos também.
        base_sel = f"tr:has-text('{ASSUNTO_CERTIDAO}'):has(td:has-text('{hoje_label}'))"
        unread = page.locator(f"tr.unread:has-text('{ASSUNTO_CERTIDAO}'):has(td:has-text('{hoje_label}'))")
        rows = unread
        if unread.count() == 0:
            rows = page.locator(base_sel)
        return rows

    while True:
        rows = buscar_rows_do_dia()
        total = rows.count()
        if total == 0:
            logger.info("[Webmail] Nenhum e-mail de hoje com o assunto alvo.")
            break

        encontrado_para_processar = False
        for i in range(total):
            row = rows.nth(i)
            # identifica o id do row (ex: rcmrowXXXX) para evitar reprocesso
            try:
                rid = row.get_attribute("id") or f"idx_{i}"
            except Exception:
                rid = f"idx_{i}"
            if rid in processed_ids:
                continue

            # abre o e-mail com duplo clique
            try:
                row.dblclick()
            except Exception:
                row.click(click_count=2)

            time.sleep(2)

            # confere se estamos vendo a mensagem certa
            try:
                if not page.get_by_text(ASSUNTO_CERTIDAO).first.is_visible():
                    # volta e segue
                    page.get_by_role("link", name="Caixa de entrada").click()
                    time.sleep(1)
                    continue
            except Exception:
                pass

            # procura o link do TJAM no corpo e abre em nova aba
            link_loc = page.get_by_role("link", name=RE_LINK_TJAM)
            if not link_loc.count():
                logger.warning("[Webmail] Link do TJAM não encontrado neste e-mail. Voltando...")
                try:
                    page.get_by_role("link", name="Caixa de entrada").click()
                except Exception:
                    page.go_back()
                time.sleep(1)
                processed_ids.add(rid)
                encontrado_para_processar = True
                continue

            try:
                with context.expect_page(timeout=15000) as nova_pg_evt:
                    link_loc.first.click()
                cert_page = nova_pg_evt.value
            except PWTimeout:
                logger.error("[Webmail] Nova aba não abriu ao clicar no link. Voltando ao inbox.")
                try:
                    page.get_by_role("link", name="Caixa de entrada").click()
                except Exception:
                    page.go_back()
                time.sleep(1)
                processed_ids.add(rid)
                encontrado_para_processar = True
                continue

            # aguarda a página da certidão carregar e salva via print (PDF)
            cert_page.wait_for_load_state("load", timeout=30000)
            # tenta extrair CNPJ do próprio link
            try:
                href = link_loc.first.get_attribute("href") or ""
            except Exception:
                href = ""
            m = re.search(r"entity\.nuCnpj=([\d\./-]+)", href)
            cnpj_num = normalizar_cnpj(m.group(1)) if m else datetime.now().strftime("%H%M%S")

            ts = datetime.now().strftime("%Y%m%d_%H%M%S")
            destino = OUTPUT_EMAIL_DIR / f"tjam_falencia_{cnpj_num}_{ts}.pdf"
            try:
                cert_page.pdf(path=str(destino), format="A4")
                logger.success(f"[Webmail] Certidão salva: {destino}")
            except Exception as e:
                # fallback: screenshot em PNG
                destino_png = OUTPUT_EMAIL_DIR / f"tjam_falencia_{cnpj_num}_{ts}.png"
                cert_page.screenshot(path=str(destino_png), full_page=True)
                logger.warning(f"[Webmail] Falha no PDF ({e}). Salvo screenshot: {destino_png}")

            # fecha a aba da certidão
            try:
                cert_page.close()
            except Exception:
                pass

            # volta para a Caixa de entrada
            try:
                page.get_by_role("link", name="Caixa de entrada").click()
            except Exception:
                page.go_back()
            time.sleep(1)

            processed_ids.add(rid)
            encontrado_para_processar = True

        if not encontrado_para_processar:
            break

    page.close()
    logger.info("[Webmail] Finalizado.")

# === Execução principal (com reprocessamento de falhas no captcha) ===
if __name__ == '__main__':
    df = pd.read_excel(PLANILHA, sheet_name=ABA)
    df[COL_CNPJ] = df[COL_CNPJ].astype(str).apply(normalizar_cnpj)

    falhas_captcha = []  # (cnpj, razao)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)

        # 1ª rodada
        for _, row in df.iterrows():
            cnpj = row[COL_CNPJ]
            razao = row[COL_RAZAO]

            try:
                logger.info(f"[Captcha] Solicitando resolução para CNPJ {cnpj}...")
                captcha_id = solicitar_captcha(API_KEY_2CAPTCHA, SITEKEY, URL_SITE)
                token = obter_resultado(API_KEY_2CAPTCHA, captcha_id)
                logger.info(f"[Captcha] Token recebido. Enviando pedido: {razao}")
                automatizar_com_token(token, cnpj, razao, context)
            except Exception as e:
                logger.error(f"[Falha Captcha] {cnpj} → {e}")
                falhas_captcha.append((cnpj, razao))
                continue

        # 2ª rodada só dos que falharam
        if falhas_captcha:
            logger.info(f"[Reprocessamento] Tentando novamente {len(falhas_captcha)} CNPJ(s).")
            for cnpj, razao in falhas_captcha:
                try:
                    captcha_id = solicitar_captcha(API_KEY_2CAPTCHA, SITEKEY, URL_SITE)
                    token = obter_resultado(API_KEY_2CAPTCHA, captcha_id)
                    automatizar_com_token(token, cnpj, razao, context)
                    logger.success(f"[Reprocessamento] {cnpj} concluído.")
                except Exception as e:
                    logger.error(f"[Reprocessamento Falhou] {cnpj} → {e}")

        # 3) Ao final, baixa as certidões do dia
        baixar_certidoes_email(context)

        context.close()
        browser.close()

        print("[3] Token recebido. Automatizando para:", razao)
        automatizar_com_token(token, cnpj, razao)
