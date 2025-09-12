from playwright.sync_api import sync_playwright
import time
import re
import traceback
from datetime import datetime
from pathlib import Path
import pandas as pd
import fitz  # PyMuPDF
from loguru import logger
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "RFB"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_STATUS = "STATUS"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

URL_RFB = "https://servicos.receitafederal.gov.br/servico/certidoes/#/home/cnpj"
REGEX_VALIDADE = r"Válida até (\d{2}/\d{2}/\d{4})"

# === Funções utilitárias ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, status: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)
    idx_status = colunas.get(COL_STATUS)

    if not idx_cnpj:
        logger.error("Coluna CNPJ não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            if idx_validade:
                row[idx_validade - 1].value = nova_data
            if idx_status:
                row[idx_status - 1].value = status
            break

    wb.save(caminho)
    wb.close()

# === Nova função robusta para preencher o CNPJ ===
def preencher_cnpj(page, cnpj: str):
    try:
        campo = page.locator("input[name='niContribuinte']").first
        campo.wait_for(state="visible", timeout=5000)
        campo.fill("")
        campo.type(cnpj)
        return True
    except Exception:
        pass

    try:
        campo = page.locator("input[placeholder='Informe o CNPJ']").first
        campo.wait_for(state="visible", timeout=5000)
        campo.fill("")
        campo.type(cnpj)
        return True
    except Exception:
        pass

    for frame in page.frames:
        try:
            campo = frame.locator("input[name='niContribuinte']").first
            if campo.count():
                campo.wait_for(state="visible", timeout=5000)
                campo.fill("")
                campo.type(cnpj)
                return True
        except Exception:
            continue

    raise RuntimeError("Campo de CNPJ não encontrado.")

# === Fluxo principal ===
def processar_certidoes():
    df = pd.read_excel(PLANILHA, sheet_name=ABA)
    df[COL_CNPJ] = df[COL_CNPJ].astype(str).apply(normalizar_cnpj)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()
        page.goto(URL_RFB)
        page.wait_for_load_state("networkidle")

        for _, row in df.iterrows():
            cnpj = row[COL_CNPJ]
            logger.info(f"Processando CNPJ {cnpj}...")

            try:
                # Preenche CNPJ
                preencher_cnpj(page, cnpj)
                time.sleep(1)

                # Clica em "+ Nova Certidão"
                page.get_by_role("button", name="+ Nova Certidão").click()
                time.sleep(5)

                # Verifica se deu erro
                if page.locator(".msg-resultado").filter(
                        has_text="Não foi possível concluir a ação"
                    ).count() > 0:
                    logger.warning(f"Erro ao processar {cnpj}")
                    salvar_valor_na_planilha(cnpj, "", "ERRO BAIXAR", PLANILHA, ABA)
                    continue

                # Se aparecer a confirmação de certidão já existente → clicar novamente
                if page.locator(".br-dialog").filter(
                        has_text="Certidão Válida Encontrada"
                    ).count() > 0:
                    page.get_by_role("button", name="+ Nova Certidão").click()
                    time.sleep(5)

                # Verifica mensagem de sucesso
                if page.locator(".msg-resultado").filter(
                        has_text="A certidão foi emitida com sucesso"
                    ).count() > 0:
                    # se o download for disparado automaticamente, prefira esperar o evento
                    download = page.wait_for_event("download", timeout=60000)
                    logger.info(f"Baixando certidão para {cnpj}...")
                    caminho_pdf = OUTPUT_DIR / f"{cnpj}_RFB_{datetime.now().strftime('%Y%m%d')}.pdf"
                    download.save_as(str(caminho_pdf))
                    validade = extrair_validade_pdf(caminho_pdf)
                    salvar_valor_na_planilha(cnpj, validade, "OK", PLANILHA, ABA)
                    logger.info(f"Certidão salva: {caminho_pdf.name}")

                # Volta para nova certidão
                if page.get_by_role("button", name="+ Nova Certidão").count():
                    page.get_by_role("button", name="+ Nova Certidão").click()
                    time.sleep(2)

            except Exception as e:
                logger.error(f"Erro no processamento do CNPJ {cnpj}: {e}")
                traceback.print_exc()
                salvar_valor_na_planilha(cnpj, "", "ERRO BAIXAR", PLANILHA, ABA)

        browser.close()

# === Execução ===
if __name__ == "__main__":
    processar_certidoes()
