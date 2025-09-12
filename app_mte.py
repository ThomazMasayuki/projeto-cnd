import re
import time
import requests
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import fitz  # PyMuPDF
from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "MTE"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_MTE = "https://eprocesso.sit.trabalho.gov.br/Entrar?ReturnUrl=%2FCertidao%2FEmitir"
TIMEOUT = 40_000
REGEX_VALIDADE = r"Válida até:\s*(\d{2}/\d{2}/\d{4})"

# credenciais
CPF_LOGIN = "02448982198"   # coloque seu CPF
SENHA_LOGIN = "15468973Tt*"

API_KEY_2CAPTCHA = "30924a201f06a3b554d7479c487fee8e"   # substitua pela sua chave real
SITEKEY_HCAPTCHA = "93b08d40-d46c-400a-ba07-6f91cda815b9"

# === Funções utilitárias ===
def normalizar_cnpj(cnpj: str) -> str:
    return re.sub(r"\D", "", str(cnpj)).zfill(14)

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

# === Funções 2Captcha para hCaptcha ===
def solicitar_hcaptcha(api_key, sitekey, url):
    payload = {
        "key": api_key,
        "method": "hcaptcha",
        "sitekey": sitekey,
        "pageurl": url,
        "json": 1
    }
    resp = requests.post("http://2captcha.com/in.php", data=payload).json()
    if resp.get("status") != 1:
        raise RuntimeError(f"[2Captcha] Falha ao enviar captcha: {resp}")
    return resp["request"]

def obter_resultado(api_key, captcha_id, tentativas=30, intervalo=7):
    for tentativa in range(tentativas):
        time.sleep(intervalo)
        resp = requests.get("http://2captcha.com/res.php", params={
            "key": api_key,
            "action": "get",
            "id": captcha_id,
            "json": 1
        }).json()
        if resp.get("status") == 1:
            return resp["request"]
        if resp.get("request") != "CAPCHA_NOT_READY":
            raise RuntimeError(f"[2Captcha] Erro inesperado: {resp}")
    raise TimeoutError("Captcha não resolvido a tempo pelo 2Captcha.")

# === Fluxo principal ===
def processar_mte():
    logger.add("execucao_mte.log", rotation="1 MB")
    logger.info("Iniciando automação GOV.BR no MTE")

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context(
            user_agent="Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                       "AppleWebKit/537.36 (KHTML, like Gecko) "
                       "Chrome/114.0.0.0 Safari/537.36",
            viewport={"width": 1920, "height": 1080}
        )
        page = context.new_page()

        try:
            logger.info(f"Acessando {URL_MTE}...")
            page.goto(URL_MTE, timeout=TIMEOUT)

            # Clica no botão "Entrar com GOV.BR"
            logger.info("Clicando em 'Entrar com GOV.BR'")
            page.click("#janela-login-gov-br a")

            # Preenche CPF
            logger.info("Preenchendo CPF...")
            campo_cpf = page.get_by_role("textbox", name="Digite seu CPF")
            campo_cpf.fill("")
            campo_cpf.type(CPF_LOGIN)
            time.sleep(1)

            # Solicita resolução do hCaptcha
            logger.info("Enviando hCaptcha para 2Captcha...")
            captcha_id = solicitar_hcaptcha(API_KEY_2CAPTCHA, SITEKEY_HCAPTCHA, page.url)
            token_resolvido = obter_resultado(API_KEY_2CAPTCHA, captcha_id)
            logger.success("Token hCaptcha resolvido com sucesso!")

            # Injeta token no campo h-captcha-response
            page.evaluate(f'''
                document.querySelector("textarea[name='h-captcha-response']").value = "{token_resolvido}";
            ''')
            page.evaluate('''
                let el = document.querySelector("textarea[name='h-captcha-response']");
                el.dispatchEvent(new Event('input', {{ bubbles: true }}));
                el.dispatchEvent(new Event('change', {{ bubbles: true }}));
            ''')

            # Clica no botão continuar
            logger.info("Clicando no botão 'Continuar'")
            btn = page.locator("button#enter-account-id")
            btn.click()

            # Verifica se senha aparece
            try:
                page.wait_for_selector("input[type='password']", timeout=15000)
                logger.info("Campo de senha carregado. Preenchendo senha...")
                page.fill("input[type='password']", SENHA_LOGIN)
                page.get_by_role("button", name="Entrar").click()
            except PWTimeout:
                logger.error("Botão 'Continuar' travou ou senha não apareceu. Abortando login.")
                return

            logger.success("Login realizado com sucesso!")

        except Exception as e:
            logger.error(f"Erro durante processo de login: {type(e).__name__}: {e}")
            traceback.print_exc()
        finally:
            context.close()
            browser.close()

    logger.info("Processo concluído.")

# === Execução ===
if __name__ == "__main__":
    processar_mte()
