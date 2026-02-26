import re
import time
import base64
import traceback
from datetime import datetime
from pathlib import Path
import requests

import pandas as pd
import fitz  # PyMuPDF
from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "CDT"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_CDT = "https://cndt-certidao.tst.jus.br/gerarCertidao.faces"
TIMEOUT = 40_000
REGEX_VALIDADE = r"Validade:\s*(\d{2}/\d{2}/\d{4})"
MAX_TENTATIVAS_CNPJ = 6
HEADLESS = False
POLLING_2CAPTCHA_SEG = 5
MAX_POLLS_2CAPTCHA = 50  

# === Utilitários ===
def normalizar_cnpj(doc: str) -> str:
    """
    Remove caracteres não numéricos e retorna o documento com padding apenas se for CNPJ.
    CPF (11 dígitos) permanece como está. 
    """
    numeros = re.sub(r"\D", "", str(doc))
    if len(numeros) in [11, 14]:
        return numeros
    raise ValueError(f"Documento inválido: {doc} → {numeros}")

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""


def salvar_valor_na_planilha(cnpj: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDÃO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if normalizar_cnpj(val) == normalizar_cnpj(cnpj):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

# === 2Captcha (image captcha) ===
def resolver_captcha_2captcha(caminho_imagem: Path, api_key: str) -> str:
    with open(caminho_imagem, "rb") as f:
        b64 = base64.b64encode(f.read()).decode("utf-8")

    payload = {
        "key": api_key,
        "method": "base64",
        "body": b64,
        "json": 1,
    }
    resp = requests.post("http://2captcha.com/in.php", data=payload, timeout=40)
    resp.raise_for_status()
    data = resp.json()
    if data.get("status") != 1:
        raise RuntimeError(f"Falha ao enfileirar captcha: {data}")

    captcha_id = data["request"]

    # Polling de resultado
    for _ in range(MAX_POLLS_2CAPTCHA):
        time.sleep(POLLING_2CAPTCHA_SEG)
        res = requests.get(
            "http://2captcha.com/res.php",
            params={"key": api_key, "action": "get", "id": captcha_id, "json": 1},
            timeout=40,
        )
        res.raise_for_status()
        ans = res.json()
        if ans.get("status") == 1:
            return ans["request"].strip()
        # Se não estiver pronto e não for CAPCHA_NOT_READY, trate como erro
        if ans.get("request") != "CAPCHA_NOT_READY":
            raise RuntimeError(f"Erro do 2Captcha: {ans}")

    raise TimeoutError("Timeout aguardando solução do 2Captcha")


# === Baixa a certidão (download ou nova aba) ===
def tentar_baixar_certidao(page, contexto, cnpj_limpo: str) -> Path | None:
    temp_path = OUTPUT_DIR / f"temp_{cnpj_limpo}.pdf"

    # 1) Tenta evento de download direto
    try:
        with page.expect_download(timeout=15000) as dl_info:
            page.get_by_role("button", name=re.compile("Emitir Certid[aã]o", re.I)).click()
        download = dl_info.value
        download.save_as(str(temp_path))
        return temp_path
    except PWTimeout:
        pass
    except Exception as e:
        logger.debug(f"Download direto falhou: {e}")

    # 2) Tenta nova aba + page.pdf (Chromium headless)
    try:
        with contexto.expect_page() as nova_aba_evento:
            page.get_by_role("button", name=re.compile("Emitir Certid[aã]o", re.I)).click()
        nova_aba = nova_aba_evento.value
        nova_aba.wait_for_load_state("networkidle", timeout=20000)
        try:
            nova_aba.pdf(path=str(temp_path), format="A4")
            nova_aba.close()
            return temp_path
        except Exception as e:
            logger.debug(f"pdf() falhou: {e}")
            evid = OUTPUT_DIR / f"screenshot_{cnpj_limpo}.png"
            nova_aba.screenshot(path=str(evid), full_page=True)
            nova_aba.close()
            return None
    except PWTimeout:
        return None


# === Fluxo principal ===
def processar_cdt():
    logger.add("execucaocdt.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")

    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ, COL_VALIDADE]].dropna(subset=[COL_CNPJ])
    cnpjs = df[COL_CNPJ].drop_duplicates().tolist()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if not API_KEY_2CAPTCHA or API_KEY_2CAPTCHA == "COLOQUE_SUA_CHAVE_AQUI":
        logger.error("Defina API_KEY_2CAPTCHA (variável de ambiente ou no código) antes de executar.")
        return

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=HEADLESS)
        context = browser.new_context(accept_downloads=True)
        page = context.new_page()            

        for cnpj in cnpjs:
            cnpj_limpo = normalizar_cnpj(cnpj)
            tentativas = 0

            while tentativas < MAX_TENTATIVAS_CNPJ:
                tentativas += 1
                try:
                    logger.info(f"Consultando CNPJ: {cnpj_limpo} (tentativa {tentativas}/{MAX_TENTATIVAS_CNPJ})")
                    page.goto(URL_CDT, timeout=TIMEOUT)
                    page.wait_for_load_state("domcontentloaded", timeout=10000)

                    # Preenche o CNPJ (campo: "Registro no Cadastro Nacional...")
                    page.get_by_role("textbox", name=re.compile("Cadastro Nacional|CNPJ", re.I)).fill(cnpj_limpo)

                    # Aguarda o captcha renderizar
                    time.sleep(1.5)

                    # Captura a imagem do captcha
                    captcha_img = page.get_by_role("img", name=re.compile("Captcha", re.I)).first
                    captcha_path = OUTPUT_DIR / f"captcha_{cnpj_limpo}.png"
                    captcha_img.screenshot(path=str(captcha_path))

                    # Resolve com 2Captcha (image captcha)
                    texto_captcha = resolver_captcha_2captcha(captcha_path, API_KEY_2CAPTCHA)
                    logger.info(f"2Captcha → '{texto_captcha}'")

                    # Preenche o captcha
                    page.get_by_role("textbox", name=re.compile("Digite os caracteres|Captcha|caracteres exibidos", re.I)).fill(texto_captcha)
                    time.sleep(0.6)

                    # Tenta emitir e obter o PDF
                    temp_pdf = tentar_baixar_certidao(page, context, cnpj_limpo)

                    if temp_pdf is None:
                        # Heurística: se há mensagem de erro de captcha, recarrega e tenta novamente
                        try:
                            erro_visivel = page.locator(
                                "text=/inv[aá]lido|c[oó]digo incorreto|captcha|caracteres/i"
                            ).first.is_visible(timeout=1000)
                        except Exception:
                            erro_visivel = False

                        if erro_visivel:
                            logger.warning("Captcha inválido/erro detectado. Recarregando captcha…")
                            try:
                                captcha_img.click()  # muitos sites recarregam a imagem ao clicar
                            except Exception:
                                pass
                            time.sleep(1.2)
                            continue  # próxima tentativa

                        logger.warning("Sem PDF nem erro claro; tentando novamente…")
                        continue

                    # Se chegou aqui, temos PDF → extrai validade e salva
                    validade = extrair_validade_pdf(temp_pdf)
                    if validade:
                        salvar_valor_na_planilha(cnpj_limpo, validade, PLANILHA, ABA)
                        validade_formatada = datetime.strptime(validade, "%d/%m/%Y").strftime("%Y%m%d")
                        destino_pdf = OUTPUT_DIR / f"cdt_{cnpj_limpo}_{validade_formatada}.pdf"
                        temp_pdf.replace(destino_pdf)
                        logger.success(f"{cnpj_limpo} → Sucesso: validade {validade}")
                    else:
                        destino_pdf = OUTPUT_DIR / f"erro_{cnpj_limpo}.pdf"
                        temp_pdf.replace(destino_pdf)
                        logger.warning(f"{cnpj_limpo} → PDF salvo, mas não foi possível extrair validade.")

                    # Sucesso → sair do loop de tentativas
                    break

                except Exception as e:
                    motivo = f"{type(e).__name__}: {e}"
                    logger.error(f"{cnpj_limpo} → ERRO: {motivo}")
                    traceback.print_exc()
                    # Loop continua até atingir o MAX_TENTATIVAS_CNPJ

            else:
                logger.error(f"{cnpj_limpo} → Excedeu o número máximo de tentativas.")

        context.close()
        browser.close()

    logger.info("Processo concluído.")


# === Execução ===
if __name__ == "__main__":
    processar_cdt()
