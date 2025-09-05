import re
import time
import traceback
from datetime import datetime
from pathlib import Path

import pandas as pd
import fitz  # PyMuPDF
from loguru import logger
from playwright.sync_api import sync_playwright, TimeoutError as PWTimeout
from openpyxl import load_workbook

# === Configurações ===
PLANILHA = Path("base_certidoes.xlsx")
ABA = "SEFAZ N CONT"
COL_CNPJ = "CNPJ"
COL_VALIDADE = "VALIDADE CERTIDÃO"
COL_RAZAO = "RAZÃO SOCIAL"
OUTPUT_DIR = Path("certidoes_baixadas") / ABA.replace(" ", "_")
URL_SEFAZ = "https://sistemas.sefaz.am.gov.br/GAE/mnt/dividaAtiva/certidaoNegativa/emitirCertidaoNegativaNaoContPortal.do"
TIMEOUT = 40_000
REGEX_VALIDADE = r"Válida até:\s*(\d{2}/\d{2}/\d{4})"

# === Funções utilitárias ===
def limpar_documento(doc: str) -> str:
    return re.sub(r"\D", "", str(doc))

def extrair_validade_pdf(caminho_pdf: Path) -> str:
    try:
        with fitz.open(caminho_pdf) as doc:
            texto = "\n".join(page.get_text() for page in doc)
        match = re.search(REGEX_VALIDADE, texto, re.IGNORECASE)
        if match:
            return match.group(1)
    except Exception as e:
        logger.warning(f"Erro ao ler validade do PDF {caminho_pdf.name}: {e}")
    return ""

def salvar_valor_na_planilha(doc: str, nova_data: str, caminho: Path, aba: str):
    wb = load_workbook(caminho)
    ws = wb[aba]
    colunas = {cell.value: idx for idx, cell in enumerate(next(ws.iter_rows(min_row=1, max_row=1)), start=1)}
    idx_cnpj = colunas.get(COL_CNPJ)
    idx_validade = colunas.get(COL_VALIDADE)

    if not idx_cnpj or not idx_validade:
        logger.error("Coluna CNPJ ou VALIDADE CERTIDAO não encontrada.")
        return

    for row in ws.iter_rows(min_row=2):
        val = str(row[idx_cnpj - 1].value)
        if limpar_documento(val) == limpar_documento(doc):
            row[idx_validade - 1].value = nova_data
            break

    wb.save(caminho)
    wb.close()

# === Função principal ===
def processar_sefaz_n_contribuinte():
    logger.add("execucao.log", rotation="1 MB")
    logger.info(f"Iniciando automação da aba: {ABA}")

    df = pd.read_excel(PLANILHA, sheet_name=ABA, dtype=str)
    df = df[[COL_RAZAO, COL_CNPJ, COL_VALIDADE]].dropna(subset=[COL_CNPJ])
    documentos = df[COL_CNPJ].drop_duplicates().tolist()

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=False)
        context = browser.new_context()
        page = context.new_page()

        for doc_bruto in documentos:
            doc = limpar_documento(doc_bruto)

            if len(doc) == 11:
                tipo = "CPF"
            elif len(doc) == 14:
                tipo = "CNPJ"
            else:
                logger.warning(f"{doc_bruto} → Documento inválido (não é CPF nem CNPJ). Pulando.")
                continue

            try:
                logger.info(f"Consultando {tipo}: {doc}")
                page.goto(URL_SEFAZ, timeout=TIMEOUT)
                page.get_by_label("CPF ou CNPJ:").fill(doc)
                page.get_by_label("CND completa").check()
                page.get_by_role("button", name="Emitir").click()
                page.wait_for_load_state("networkidle", timeout=15000)

                # Gera o conteúdo do PDF diretamente na memória
                pdf_bytes = page.pdf(format="A4")

                # Salva temporariamente na memória para leitura com PyMuPDF
                temp_path = OUTPUT_DIR / f"temp_{doc}.pdf"
                with open(temp_path, "wb") as f:
                    f.write(pdf_bytes)
                    
                validade = extrair_validade_pdf(temp_path)

                if validade:
                    salvar_valor_na_planilha(doc, validade, PLANILHA, ABA)
                    validade_formatada = datetime.strptime(validade, "%d/%m/%Y").strftime("%Y%m%d")
                    nome_arquivo = f"sefaz_n_contribuinte_{doc}_{validade_formatada}.pdf"
                    destino_pdf = OUTPUT_DIR / nome_arquivo

                    # Agora salva com o nome definitivo
                    with open(destino_pdf, "wb") as f:
                        f.write(pdf_bytes)

                    # Remove temporário
                    temp_path.unlink(missing_ok=True)

                    logger.success(f"{doc} → Sucesso: validade {validade}")
                else:
                    destino_pdf = OUTPUT_DIR / f"erro_{doc}.pdf"
                    with open(destino_pdf, "wb") as f:
                        f.write(pdf_bytes)
                    logger.warning(f"{doc} → Não foi possível extrair validade.")

            except Exception as e:
                motivo = f"{type(e).__name__}: {e}"
                logger.error(f"{doc} → ERRO: {motivo}")
                traceback.print_exc()

        context.close()
        browser.close()

    logger.info("Processo concluído.")

# === Execução ===
if __name__ == "__main__":
    processar_sefaz_n_contribuinte()